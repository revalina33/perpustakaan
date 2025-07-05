from flask import Flask, render_template, request, jsonify, send_from_directory, abort
import os
import zipfile
import pandas as pd
import pyexcel as pe
from werkzeug.utils import secure_filename
import re
import calendar
from datetime import datetime, timedelta, date
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import xlrd
from bs4 import BeautifulSoup
from typing import Dict, List, Optional, Tuple
import numpy as np
from dateutil.parser import parse

# ==============================================
# CONFIGURATION
# ==============================================
app = Flask(__name__)
app.config.update({
    'UPLOAD_FOLDER': 'uploads',
    'ALLOWED_EXTENSIONS': {'xlsx', 'xls', 'csv', 'html'},
    'MAX_CONTENT_LENGTH': 10 * 1024 * 1024  # 10MB
})

# kategori data
CATEGORIES = [
    'TK/Paud', 'SD', 'SLTP', 'SLTA', 'Mahasiswa', 
    'GURU/DOSEN', 'PEG/KARY.', 'UMUM', 'Grup/Rombongan', 'data tidak lengkap'
]


os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Logging Configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ==============================================
# UTILITY CLASSES
# ==============================================
class FileValidator:
    """Handles file validation operations"""
    
    @staticmethod
    def allowed_file(filename: str) -> bool:
        """Check if file extension is allowed"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

    @staticmethod
    def is_html_file(filepath: str) -> bool:
        """Check if file is HTML disguised as Excel"""
        try:
            with open(filepath, 'rb') as f:
                start = f.read(512).lower()
                return b'<html' in start or b'<!doctype html' in start
        except Exception as e:
            logger.error(f"Error checking HTML file: {str(e)}")
            return False

    @staticmethod
    def is_valid_excel(filepath: str) -> bool:
        """Validate Excel file based on format"""
        try:
            if filepath.endswith('.csv'):
                try:
                    pd.read_csv(filepath, nrows=1)
                    return True
                except Exception:
                    return False
            
            elif filepath.endswith('.xlsx'):
                try:
                    with zipfile.ZipFile(filepath) as zf:
                        return any(name.startswith('xl/') for name in zf.namelist())
                except Exception:
                    return False
            
            elif filepath.endswith('.xls'):
                try:
                    with open(filepath, 'rb') as f:
                        header = f.read(8)
                        if header == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
                            return True
                        f.seek(0)
                        start = f.read(512).lower()
                        if b'<html' in start:
                            return True
                    return False
                except Exception:
                    return False
            
            return False
        except Exception as e:
            logger.error(f"Error validating file: {str(e)}")
            return False


class FileReader:
    """Handles file reading operations for different file types"""
    
    @staticmethod
    def read_html_file(filepath: str) -> pd.DataFrame:
        """Read HTML file and extract tables"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            soup = BeautifulSoup(content, 'html.parser')
            tables = soup.find_all('table')
            
            if not tables:
                raise ValueError("No tables found in HTML file")
            
            for table in tables:
                try:
                    df = pd.read_html(str(table))[0]
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    df.columns = [str(col).strip() for col in df.columns]
                    if len(df.columns) > 1 and len(df) > 0:
                        return df
                except Exception as e:
                    logger.warning(f"Skipping table due to error: {str(e)}")
                    continue
            
            raise ValueError("No valid tables found in HTML file")
        except Exception as e:
            raise Exception(f"Failed to read HTML file: {str(e)}")

    @staticmethod
    def read_csv_file(filepath: str) -> pd.DataFrame:
        """Read CSV file with multiple attempts"""
        try:
            # Try reading with header detection
            try:
                df = pd.read_csv(filepath)
                if len(df.columns) < 2:  # If likely no header
                    df = pd.read_csv(filepath, header=None)
            except Exception:
                df = pd.read_csv(filepath, header=None, encoding='latin1')
            return df
        except Exception as e:
            raise Exception(f"CSV read failed: {str(e)}")

    @staticmethod
    def read_excel_file(filepath: str, filename: str) -> pd.DataFrame:
        """Read Excel file with multiple fallback methods and reliable date detection"""
        try:
            # memeriksa file html yang menyamar sebagai excel
            if FileValidator.is_html_file(filepath):
                return FileReader.read_html_file(filepath)

            if filename.endswith('.csv'):
                return FileReader.read_csv_file(filepath)

            df = None
            last_error = None

            # mengunakan berbagai parameter
            read_attempts = [
                {'engine': 'openpyxl', 'header': 0},
                {'engine': 'openpyxl', 'header': None},
                {'engine': 'xlrd', 'header': 0},
                {'engine': 'xlrd', 'header': None}
            ]

            for attempt in read_attempts:
                try:
                    df = pd.read_excel(filepath, **attempt)
                    if not df.empty and len(df.columns) > 1:
                        break
                except Exception as e:
                    last_error = e
                    continue

    
            if (df is None or df.empty) and filename.endswith('.xls'):
                try:
                    logger.info("Trying pyexcel as fallback for .xls file")
                    sheet = pe.get_sheet(file_name=filepath)
                    df = pd.DataFrame(sheet.to_array())
                    if not df.empty:
                        if any(isinstance(x, str) for x in df.iloc[0]):
                            df.columns = df.iloc[0]
                            df = df[1:]
                except Exception as e:
                    last_error = e
                    logger.error(f"Pyexcel fallback failed: {str(e)}")

            if df is None or df.empty:
                raise Exception(f"All read methods failed. Last error: {str(last_error)}")

            # Clean dataframe
            df = df.dropna(how='all').dropna(axis=1, how='all')
            df.columns = [str(col).strip() for col in df.columns]

            # kolom unnamed 1 sebagai kolom tanggal
            if 'Unnamed: 1' in df.columns:
                df['Tanggal'] = df['Unnamed: 1'].apply(DataProcessor.parse_custom_date)
            else:
                df['Tanggal'] = None
                logger.warning("Kolom 'Unnamed: 1' tidak ditemukan, tidak bisa buat kolom Tanggal")

            # Buat kolom Hari dari tanggal
            df['Hari'] = df['Tanggal'].apply(lambda x: x.day if pd.notna(x) else None)

            # Tambahkan kolom nomor urut
            df.insert(0, 'No.', range(1, len(df) + 1))

            return df

        except Exception as e:
            raise Exception(f"Failed to read file: {str(e)}")


class DataProcessor:
    """Handles data processing and analysis"""
    
    @staticmethod
    def parse_custom_date(date_str) -> Optional[date]:
        """Parse text date with multiple fallback methods"""
        if pd.isna(date_str) or str(date_str).strip() == '':
            return None

        # Handle if already a date object
        if isinstance(date_str, (date, pd.Timestamp)):
            return date_str

        date_str = str(date_str).strip()

        # menangani data yang berbentuk numerik/angka
        if isinstance(date_str, (int, float)) or date_str.replace('.', '', 1).isdigit():
            try:
                return xlrd.xldate.xldate_as_datetime(float(date_str), 0).date()
            except:
                pass

        # mengganti . di jam pada kolom tanggal menjadi : (e.g., 09.31 becomes 09:31)
        date_str = re.sub(r'(\d{1,2})\.(\d{1,2})', r'\1:\2', date_str)

        date_formats = [
            '%m/%d/%Y %I:%M:%S %p',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %I:%M %p',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y',
            '%d/%m/%Y %I:%M:%S %p',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %I:%M %p',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d %b %Y',
            '%d %B %Y',
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        # Fallback using dateutil
        try:
            parsed = parse(date_str, dayfirst=True)
            return parsed.date()
        except:
            return None
        
    @staticmethod
    def count_group_rows(row: pd.Series) -> Dict[str, int]:
        """Count group members matching Excel approach"""
        counts = {'L': 0, 'P': 0, 'U': 0}
        
        # memeriksa apakah kolom ini sama (sama seperti klasifikasi)
        name = str(row.get('Name', row.get('Unnamed: 5', ''))).upper()
        nomor_kunjungan = str(row.get('No. Kunjungan', row.get('Unnamed: 4', ''))).upper()
        
        if not ('GRP' in nomor_kunjungan or 'GRP' in name or 
                any(x.startswith('GRP') for x in [nomor_kunjungan, name])):
            return counts
        
        # jika benar grup maka hitung dari nama/deskripsi
        text = ' '.join(str(x) for x in row if pd.notna(x)).upper()
        
        # melihat angka/jumlah yang diikuti gender atau berada dalam kolom gender (misalnya 25 peremuan 10 laki - laki angkanya yang diambil)
        matches = re.findall(r'(\d+)\s*(LAKI|PRIA|MALE|PEREMPUAN|WANITA|FEMALE)', text)
        
        for number, gender in matches:
            if gender in ['LAKI', 'PRIA', 'MALE']:
                counts['L'] += int(number)
            elif gender in ['PEREMPUAN', 'WANITA', 'FEMALE']:
                counts['P'] += int(number)
        
        # jika masuk kategori grup tapi tidak terdapat jenis kelamin maka akan masuk uknown
        if counts['L'] == 0 and counts['P'] == 0:
            counts['U'] = 1
            
        return counts

    @staticmethod
    def combine_group_rows(df: pd.DataFrame) -> pd.DataFrame:
        combined_rows = []
        skip_next = False

        # Temukan kolom nomor kunjungan
        nomor_kunjungan_col = None
        for col in df.columns:
            if 'nomor' in str(col).lower() or 'kunjungan' in str(col).lower() or str(col).lower() == 'unnamed: 4':
                nomor_kunjungan_col = col
                break

        for i in range(len(df)):
            if skip_next:
                skip_next = False
                continue

            row = df.iloc[i]
            is_grp = False

            if nomor_kunjungan_col and nomor_kunjungan_col in row:
                val = str(row[nomor_kunjungan_col]).strip().upper()
                if re.match(r'^GRP\s?\d+', val):
                    is_grp = True

            if is_grp and i + 1 < len(df):
                next_row = df.iloc[i + 1]
                val_next = str(next_row.get(nomor_kunjungan_col, '')).strip().upper()

                #  Hanya gabung jika baris berikutnya kosong nomor kunjungan
                if val_next in ['', 'NAN']:
                    combined = row.copy()

                    for col in df.columns:
                        if col == 'Tanggal' or pd.api.types.is_datetime64_any_dtype(df[col]):
                            continue
                        val1 = str(row.get(col, '')).strip() if pd.notna(row.get(col)) else ''
                        val2 = str(next_row.get(col, '')).strip() if pd.notna(next_row.get(col)) else ''
                        combined[col] = f"{val1} {val2}".strip()

                    combined_rows.append(combined)
                    skip_next = True
                else:
                    # Tidak digabung karena baris berikutnya punya nomor kunjungan valid
                    combined_rows.append(row)
            else:
                combined_rows.append(row)

        return pd.DataFrame(combined_rows, columns=df.columns)


    @staticmethod
    def classify_visitor_combined(row: pd.Series) -> Optional[str]:
        """Classify visitor to match Excel formula exactly"""
        # mengambil semua baris yang sesuai
        name = str(row.get('Name', row.get('Unnamed: 5', ''))).upper()
        pekerjaan = str(row.get('Pekerjaan', row.get('Unnamed: 7', ''))).upper()
        pendidikan = str(row.get('Pendidikan', row.get('Unnamed: 8', ''))).upper()
        nomor_kunjungan = str(row.get('No. Kunjungan', row.get('Unnamed: 4', ''))).upper()
        other_fields = str(row.get('Unnamed: 6', '')).upper()  # Gender column

        # mengambungkan baris untuk dianalisis lebih lanjut
        combined = f"{name}{pekerjaan}{pendidikan}{other_fields}{nomor_kunjungan}"
        clean_pddk = re.sub(r'-', '', pendidikan)

        # 1. Grup/Rombongan 
        if ('GRP' in nomor_kunjungan or 
            'GRP' in name or 
            any(x.startswith('GRP') for x in [nomor_kunjungan, name])):
            return "Grup/Rombongan"

        # 2. UMUM check - hanya jika kata 'UMUM' muncul di nama & tidak ada pendidikan/pekerjaan yang lebih jelas
        if ('UMUM' in name and 
            not any(x in pendidikan for x in ['SD', 'SMP', 'SLTP', 'SMA', 'SLTA', 'TK', 'PAUD']) and 
            not any(x in pekerjaan for x in ['GURU', 'DOSEN', 'PELAJAR', 'MAHASISWA', 'PNS', 'SWASTA', 'WIRASWASTA'])):
            return "UMUM"

        # 3. Employment types
        if any(x in pekerjaan for x in ['WIRASWASTA', 'WIRAUSAHA', 'PEGAWAI NEGERI', 'PNS', 'PEGAWAI SWASTA', 'SWASTA']):
            return "PEG/KARY."

        if any(x in pekerjaan for x in ['IBU RUMAH TANGGA', 'IRT']):
            return "UMUM"

        if any(x in pekerjaan for x in ['GURU', 'DOSEN', 'PENGAJAR', 'LECTURER']):
            return "GURU/DOSEN"

        # 4. Mahasiswa
        if ('MAHASISWA' in combined or 
            any(x in clean_pddk for x in ['D1', 'D2', 'D3', 'S1', 'S2', 'S3', 'SARJANA'])):
            return "Mahasiswa"

        # 5. level pendidikan
        if any(x in combined for x in ['SMA', 'SMK', 'SLTA']):
            return "SLTA"

        if any(x in combined for x in ['SMP', 'SLTP']):
            return "SLTP"

        if 'SD' in combined:
            return "SD"

        if any(x in combined for x in ['TK', 'PAUD', 'TAMAN KANAK']):
            return "TK/Paud"

        # 6. memeriksa data kosong/tidak lengkap
        if not name.strip() and not pekerjaan.strip() and not pendidikan.strip():
            return "data tidak lengkap"

        return "data tidak lengkap"



    @staticmethod
    def extract_gender(row: pd.Series) -> Optional[str]:
        """Extract gender from row data"""
        for col in row.index:
            col_str = str(col).strip().lower()
            val = str(row[col]).upper().strip()

            if 'jenis kelamin' in col_str or 'gender' in col_str or 'jk' == col_str or 'sex' in col_str or col_str.startswith('unnamed: 6'):
                if val in ['L', 'LAKI-LAKI', 'PRIA', 'MALE']:
                    return "L"
                elif val in ['P', 'PEREMPUAN', 'WANITA', 'FEMALE']:
                    return "P"
        
        # If no explicit gender found, fallback to text scanning
        text = ' '.join(str(x) for x in row if pd.notna(x)).upper()
        
        if not text.strip():
            return None
        
        clean_text = re.sub(r'[^A-Z\s]', '', text)
        clean_text = ' '.join(clean_text.split())

        female_patterns = [r'\bPEREMPUAN\b', r'\bWANITA\b', r'\bP\b(?!\w)', r'\bPR\b', r'\bCEWEK\b', r'\bFEMALE\b']
        male_patterns = [r'\bLAKI[-\s]?LAKI\b', r'\bLAKI\b', r'\bL\b(?!\w)', r'\bPRIA\b', r'\bCOWOK\b', r'\bMALE\b']

        for pattern in female_patterns:
            if re.search(pattern, clean_text):
                return "P"
        for pattern in male_patterns:
            if re.search(pattern, clean_text):
                return "L"
        return None

    @staticmethod
    def determine_month_year(df_raw: pd.DataFrame) -> Tuple[int, int, int]:
        """Determine month and year from dataframe with fallback"""
        if 'Tanggal' in df_raw.columns:
            valid_dates = df_raw['Tanggal'].dropna()
            if not valid_dates.empty:
                first_date = valid_dates.iloc[0]
                month = first_date.month
                year = first_date.year
                days_in_month = calendar.monthrange(year, month)[1]
                return month, year, days_in_month
        
        # Fallback to current month if no valid dates found
        now = datetime.now()
        month = now.month
        year = now.year
        days_in_month = calendar.monthrange(year, month)[1]
        return month, year, days_in_month

    @staticmethod 
    def init_report_structures(days_in_month: int) -> Tuple[pd.DataFrame, Dict[str, int]]:
        """Initialize report structures with proper columns"""
        # Daily report columns
        columns = ['TANGGAL']
        for cat in CATEGORIES:
            columns.extend([f'{cat}_L', f'{cat}_P', f'{cat}_U'])
        
        # Initialize DataFrame
        rekap_harian = pd.DataFrame(
            0,
            index=range(1, days_in_month + 1),
            columns=columns
        )
        rekap_harian['TANGGAL'] = range(1, days_in_month + 1)
        
        # Initialize totals dictionary
        totals = {}
        for cat in CATEGORIES:
            totals.update({
                f'{cat}_L': 0,
                f'{cat}_P': 0, 
                f'{cat}_U': 0
            })
        
        return rekap_harian, totals

    @staticmethod
    def get_day_from_row(row: pd.Series, days_in_month: int, idx: int) -> int:
        """Get day from row with proper validation and fallback"""
        if "Tanggal" in row.index and pd.notna(row["Tanggal"]):
            try:
                return row["Tanggal"].day
            except Exception:
                pass

        if "Hari" in row.index and pd.notna(row["Hari"]):
            try:
                day = int(row["Hari"])
                if 1 <= day <= days_in_month:
                    return day
            except Exception:
                pass

        return (idx % days_in_month) + 1

    @staticmethod
    def process_dataframe(df_raw: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, int], Dict[str, int]]:
        """Process dataframe with enhanced validation and error handling"""
        import numpy as np
        
        # Create a copy to avoid modifying the original dataframe
        df = df_raw.copy()

        # 1. Handle date column
        if 'Tanggal' not in df.columns:
            df['Tanggal'] = None

        # 2. Parse dates with error handling
        df['Tanggal'] = df['Tanggal'].apply(
            lambda x: DataProcessor.parse_custom_date(x) if pd.notna(x) else None
        )

        # 3. Filter only valid dates and extract day
        df = df[df['Tanggal'].notna()].copy()
        df['Hari'] = df['Tanggal'].dt.day

        # 4. Extract gender
        df['Jenis Kelamin'] = df.apply(DataProcessor.extract_gender, axis=1)

        # 5. Handle 'Name' column if not named properly
        if 'Name' not in df.columns and 'Unnamed: 5' in df.columns:
            df['Name'] = df['Unnamed: 5'].astype(str).str.strip()

        # 6. Classify visitors
        def classify_row(row):
            if row.get('Klasifikasi') == 'data tidak lengkap':
                return 'data tidak lengkap'
            return DataProcessor.classify_visitor_combined(row)

        df['Klasifikasi'] = df.apply(classify_row, axis=1)

        # 7. Detect incomplete data
        unknown_gender_names = []
        incomplete_mask = (
            (df['Jenis Kelamin'].isna()) |
            (df['Jenis Kelamin'] == 'U') |
            (df['Klasifikasi'] == 'data tidak lengkap')
        )

        for idx, row in df[incomplete_mask].iterrows():
            name = ''
            if 'Name' in row and pd.notna(row['Name']):
                name = str(row['Name']).strip()
            elif 'Unnamed: 5' in row and pd.notna(row['Unnamed: 5']):
                name = str(row['Unnamed: 5']).strip()

            if name:
                unknown_gender_names.append(name)
            df.at[idx, 'Klasifikasi'] = 'data tidak lengkap'

        # 8. Remove invalid classifications
        df = df[df['Klasifikasi'].notna()]

        # 9. Get month/year and prepare rekap structures
        month, year, days_in_month = DataProcessor.determine_month_year(df)
        rekap_harian, totals = DataProcessor.init_report_structures(days_in_month)

        # 10. Rekap loop
        for _, row in df.iterrows():
            kategori = row['Klasifikasi']
            gender = row['Jenis Kelamin'] or 'U'
            try:
                day = int(row['Hari'])
                if day < 1 or day > days_in_month:
                    day = 1
            except:
                day = 1

            if kategori == "Grup/Rombongan":
                members = DataProcessor.count_group_rows(row)
                if sum(members.values()) > 0:
                    for g, count in members.items():
                        if count > 0:
                            col_name = f"{kategori}_{g}"
                            rekap_harian.at[day, col_name] += count
                            totals[col_name] += count
            elif kategori in CATEGORIES:
                col_name = f"{kategori}_{gender}"
                if pd.isna(rekap_harian.at[day, col_name]):
                    rekap_harian.at[day, col_name] = 1
                else:
                    rekap_harian.at[day, col_name] += 1
                totals[col_name] += 1

        # 11. Tambah total baris
        rekap_harian['JUMLAH'] = rekap_harian.iloc[:, 1:].sum(axis=1)
        total_row = rekap_harian.iloc[:, 1:].sum()
        total_row['TANGGAL'] = 'Jumlah'
        rekap_harian = pd.concat([rekap_harian, total_row.to_frame().T], ignore_index=True)

        # 12. Return
        return df, rekap_harian, totals, {
            'month': month,
            'year': year,
            'days_in_month': days_in_month,
            'unknown_gender_names': unknown_gender_names
        }



class ExcelGenerator:
    """Handles Excel file generation with formatting"""
    
    @staticmethod
    def apply_worksheet_styling(worksheet, df, header_fill, header_font, header_alignment):
        """Apply common styling to worksheet"""
        # Format header
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Format serial number column
        if 'No.' in df.columns:
            no_col_idx = df.columns.get_loc('No.') + 1
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                cell = row[no_col_idx-1]
                cell.alignment = Alignment(horizontal='center')
        
        # Set column width
        for column in worksheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in column),
                default=0
            )
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 30)

    @staticmethod
    def highlight_unknown_gender(worksheet, df):
        """Highlight unknown gender cells"""
        if 'Jenis Kelamin' in df.columns:
            col_idx = df.columns.get_loc('Jenis Kelamin') + 1
            unknown_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            unknown_font = Font(color='9C0006', italic=True)
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                cell = row[col_idx-1]
                if cell.value is None or str(cell.value).strip() == '':
                    cell.value = 'Tidak Diketahui'
                    cell.fill = unknown_fill
                    cell.font = unknown_font

    @staticmethod
    def highlight_group_rows(worksheet, df):
        """Highlight group rows in green"""
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        for col_idx, col_name in enumerate(df.columns, 1):
            col_text = str(col_name).lower()
            if any(keyword in col_text for keyword in ['nomor', 'no', 'number', 'kunjungan', 'visit']):
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    cell = row[col_idx-1]
                    if cell.value and re.match(r'^GRP\d*', str(cell.value).upper()):
                        for cell_in_row in row:
                            cell_in_row.fill = green_fill
                break

    @staticmethod
    def highlight_empty_dates(worksheet, df):
        """Highlight empty or invalid dates"""
        if 'Tanggal' in df.columns:
            date_col_idx = df.columns.get_loc('Tanggal') + 1
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                date_cell = row[date_col_idx-1]
                if date_cell.value is None or str(date_cell.value).strip() == '':
                    # Check if this is SD/SMA
                    klasifikasi = None
                    for cell in row:
                        if str(cell.value).upper() in ['SD', 'SMA', 'SLTA']:
                            klasifikasi = str(cell.value).upper()
                            break
                    
                    if klasifikasi in ['SD', 'SMA', 'SLTA']:
                        for cell_in_row in row:
                            cell_in_row.fill = yellow_fill
                    else:
                        for cell_in_row in row:
                            cell_in_row.fill = red_fill

    @staticmethod
    def create_summary_sheet(workbook, totals):
        """Create summary sheet with totals"""
        worksheet_summary = workbook.create_sheet(title='Rekap Total')
        start_row = 4
        
        # Add title
        worksheet_summary.merge_cells('A1:G1')
        title_cell = worksheet_summary['A1']
        title_cell.value = 'REKAPITULASI PENGUNJUNG'
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ["", "Kategori", "Laki-laki", "Perempuan", "Tidak Diketahui", "Total"]
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet_summary.cell(row=start_row, column=col_num, value=header)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # Add data rows
        for i, cat in enumerate(CATEGORIES, start=1):
            row_num = start_row + i
            
            worksheet_summary.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal='center')
            worksheet_summary.cell(row=row_num, column=2, value=cat)
            worksheet_summary.cell(row=row_num, column=3, value=totals.get(f'{cat}_L', 0))
            worksheet_summary.cell(row=row_num, column=4, value=totals.get(f'{cat}_P', 0))
            worksheet_summary.cell(row=row_num, column=5, value=totals.get(f'{cat}_U', 0))
            worksheet_summary.cell(row=row_num, column=6, value=f"=SUM(C{row_num}:E{row_num})")
            
            for col in range(1, 7):
                cell = worksheet_summary.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                if col > 2:
                    cell.number_format = '#,##0'
        
        # Add totals row
        total_row = start_row + len(CATEGORIES) + 1
        worksheet_summary.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        
        for col in range(3, 7):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}{start_row+1}:{col_letter}{total_row-1})" if col < 6 else f"=SUM(F{start_row+1}:F{total_row-1})"
            cell = worksheet_summary.cell(row=total_row, column=col, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0'
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        # Set column widths
        worksheet_summary.column_dimensions['A'].width = 5
        worksheet_summary.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F']:
            worksheet_summary.column_dimensions[col].width = 12

    @staticmethod
    def create_incomplete_data_sheet(workbook, df_raw):
        """Create a sheet listing incomplete data entries"""
        worksheet = workbook.create_sheet(title='Data Tidak Lengkap')

        name_col = 'Name' if 'Name' in df_raw.columns else 'Unnamed: 5' if 'Unnamed: 5' in df_raw.columns else None
        gender_col = 'Unnamed: 6'
        job_col = 'Unnamed: 7'
        edu_col = 'Unnamed: 8'

        if name_col is None:
            worksheet['A1'] = "Tidak ada data tidak lengkap"
            return

        # Label yang tidak dianggap nama orang
        label_kategori = ['tk', 'paud', 'sd', 'smp', 'sltp', 'sma', 'slta',
                        'mahasiswa', 'umum', 'guru', 'dosen', 'pelajar', 'lain-lain']

        def is_real_name(nama):
            if pd.isna(nama): return False
            nama = str(nama).lower().strip()
            return not any(k in nama for k in label_kategori)

        mask = df_raw['Klasifikasi'] == 'data tidak lengkap'
        incomplete_data = df_raw.loc[mask].copy()


        # Judul
        worksheet.merge_cells('A1:E1')
        title_cell = worksheet['A1']
        title_cell.value = 'DATA TIDAK LENGKAP'
        title_cell.font = Font(size=12, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Header
        headers = ["No.", "Tanggal", "Nama", "Jenis Kelamin", "Keterangan"]
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col_num, value=header)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

        # Data rows
        for i, (_, row) in enumerate(incomplete_data.iterrows(), start=1):
            row_num = 3 + i
            worksheet.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal='center')

            tanggal = row.get('Tanggal')
            tanggal_val = tanggal.strftime('%d-%m-%Y') if pd.notna(tanggal) and hasattr(tanggal, 'strftime') else str(tanggal) if pd.notna(tanggal) else 'Tidak Diketahui'
            worksheet.cell(row=row_num, column=2, value=tanggal_val)

            worksheet.cell(row=row_num, column=3, value=row[name_col])
            worksheet.cell(row=row_num, column=4, value=row[gender_col])
            worksheet.cell(row=row_num, column=5, value='Pekerjaan dan Pendidikan kosong')

            for col in range(1, 6):
                cell = worksheet.cell(row=row_num, column=col)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        # Lebar kolom
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 35

        # Total
        count_row = 4 + len(incomplete_data)
        worksheet.cell(row=count_row, column=1, value="Total:").font = Font(bold=True)
        worksheet.cell(row=count_row, column=2, value=len(incomplete_data)).font = Font(bold=True)


    @staticmethod
    def generate_excel_output(df_raw: pd.DataFrame, rekap_harian: pd.DataFrame, totals: Dict[str, int], output_path: str) -> bool:
        """Generate Excel output with enhanced formatting"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Raw Data
                df_raw.to_excel(writer, sheet_name='Data Mentah', index=False)
                workbook = writer.book
                worksheet_raw = writer.sheets['Data Mentah']
                
                # Apply styling to raw data sheet
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                header_alignment = Alignment(horizontal='center')
                
                ExcelGenerator.apply_worksheet_styling(worksheet_raw, df_raw, header_fill, header_font, header_alignment)
                ExcelGenerator.highlight_unknown_gender(worksheet_raw, df_raw)
                ExcelGenerator.highlight_group_rows(worksheet_raw, df_raw)
                ExcelGenerator.highlight_empty_dates(worksheet_raw, df_raw)
                
                # Sheet 2: Daily Summary
                rekap_harian.to_excel(writer, sheet_name='Rekap Harian', index=False)
                worksheet_daily = writer.sheets['Rekap Harian']
                
                # Format header for daily summary
                daily_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                for row in worksheet_daily.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = daily_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Set column width for daily summary
                for column in worksheet_daily.columns:
                    worksheet_daily.column_dimensions[get_column_letter(column[0].column)].width = 12
                
                # Sheet 3: Summary
                ExcelGenerator.create_summary_sheet(workbook, totals)
                
                # Sheet 4: Incomplete Data
                ExcelGenerator.create_incomplete_data_sheet(workbook, df_raw)
                
                # Reorder sheets
                workbook._sheets.sort(key=lambda ws: ['Data Mentah', 'Rekap Harian', 'Rekap Total', 'Data Tidak Lengkap'].index(ws.title))
                
            return True
        except Exception as e:
            logger.error(f"Error generating Excel output: {str(e)}")
            return False

# ==============================================
# FLASK ROUTES
# ==============================================
@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():  
    """Handle file upload and processing"""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file part"})

    file = request.files['file']

    if file.filename == '':
        return jsonify({"success": False, "error": "No selected file"})

    if not FileValidator.allowed_file(file.filename):
        return jsonify({
            "success": False, 
            "error": "File type not allowed. Only Excel (.xlsx, .xls), CSV, or HTML files are accepted"
        })

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        # Ensure upload directory exists
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # Save the file
        file.save(filepath)
        logger.info(f"File saved to: {filepath}")

        # Check if file is valid
        if not FileValidator.is_valid_excel(filepath) and not FileValidator.is_html_file(filepath):
            os.remove(filepath)
            return jsonify({
                "success": False, 
                "error": "Invalid file format. Please upload a valid Excel, CSV, or HTML file."
            })

        try:
            df_raw = FileReader.read_excel_file(filepath, filename)
            df_raw = DataProcessor.combine_group_rows(df_raw)
        except Exception as e:
            os.remove(filepath)
            logger.error(f"Error reading file: {str(e)}", exc_info=True)
            return jsonify({
                "success": False,
                "error": f"Failed to read file. Error details: {str(e)}\n"
                         "Possible solutions:\n"
                         "1. Try opening and saving the file in Microsoft Excel first\n"
                         "2. Convert the file to .xlsx format\n"
                         "3. Ensure the file is not password protected"
            })

        if df_raw.empty:
            os.remove(filepath)
            return jsonify({
                "success": False, 
                "error": "The file is empty or couldn't be read properly"
            })

    
        # Process data
        try:
            unknown_entries = []
            if 'Tanggal' in df_raw.columns:
                name_col = 'Name' if 'Name' in df_raw.columns else 'Unnamed: 5'
                gender_col = 'Unnamed: 6'
                job_col = 'Unnamed: 7'
                edu_col = 'Unnamed: 8'

                kategori_label = [
                    'tk', 'paud', 'sd', 'smp', 'sltp', 'sma', 'slta',
                    'mahasiswa', 'umum', 'guru', 'dosen', 'pelajar', 'lain-lain',
                    'umum laki-laki', 'umum perempuan'
                ]

                def is_empty(val):
                    if pd.isna(val): return True
                    val = str(val).strip().lower()
                    return val in ['', '-', 'pelajar', 'lain-lain']

                def is_real_name(name):
                    if pd.isna(name): return False
                    name = str(name).lower().strip()
                    return not any(k in name for k in kategori_label)

                # Mask untuk gender tidak diketahui
                gender_series = df_raw[gender_col].astype(str).str.strip().str.lower()
                mask_unknown_gender = gender_series.isin(['', 'u'])

                # Mask pekerjaan dan pendidikan kosong
                mask_missing_job_edu = (
                    df_raw[job_col].apply(is_empty) &
                    df_raw[edu_col].apply(is_empty) &
                    ~mask_unknown_gender &
                    df_raw[name_col].apply(is_real_name)
                )

                combined_mask = mask_unknown_gender | mask_missing_job_edu

                unknown_entries = df_raw.loc[combined_mask, [name_col, 'Tanggal']].dropna(subset=[name_col]).to_dict('records')
                df_raw.loc[combined_mask, 'Klasifikasi'] = 'data tidak lengkap'



            df_raw, rekap_harian, totals, date_info = DataProcessor.process_dataframe(df_raw)
        except Exception as e:
            os.remove(filepath)
            logger.error(f"Error processing data: {str(e)}", exc_info=True)
            return jsonify({
                "success": False,
                "error": f"Error processing data: {str(e)}"
            })

        # Generate output file
        output_filename = 'hasil_' + filename
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        if not ExcelGenerator.generate_excel_output(df_raw, rekap_harian, totals, output_path):
            os.remove(filepath)
            return jsonify({
                "success": False, 
                "error": "Failed to generate output Excel file"
            })

        # Prepare results
        total_male = sum(totals.get(f'{cat}_L', 0) for cat in CATEGORIES)
        total_female = sum(totals.get(f'{cat}_P', 0) for cat in CATEGORIES)
        total_unknown = sum(totals.get(f'{cat}_U', 0) for cat in CATEGORIES)

        details = [{
            "category": cat,
            "male": totals.get(f'{cat}_L', 0),
            "female": totals.get(f'{cat}_P', 0),
            "unknown": totals.get(f'{cat}_U', 0),
            "total": totals.get(f'{cat}_L', 0) + totals.get(f'{cat}_P', 0) + totals.get(f'{cat}_U', 0)
        } for cat in CATEGORIES]

        # Format tanggal untuk response
        for entry in unknown_entries:
            if isinstance(entry['Tanggal'], (datetime, pd.Timestamp)):
                entry['Tanggal'] = entry['Tanggal'].strftime('%Y-%m-%d')

        results = {
            "filename": output_filename,
            "total": total_male + total_female + total_unknown,
            "male": total_male,
            "female": total_female,
            "unknown_gender": total_unknown,
            "unknown_entries": unknown_entries,  # Sekarang berisi nama dan tanggal
            "details": details,
            "days_in_month": date_info['days_in_month'],
            "month": date_info['month'],
            "year": date_info['year']
        }

        return jsonify({"success": True, "result": results})

    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        logger.error(f"Error processing file: {str(e)}", exc_info=True)
        return jsonify({
            "success": False,
            "error": f"An error occurred: {str(e)}"
        })

@app.route('/download/<filename>')
def download_file(filename):
    """Handle file downloads"""
    filename = secure_filename(filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        abort(404)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


# ==============================================
# MAIN EXECUTION
# ==============================================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)